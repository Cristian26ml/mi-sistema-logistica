from openpyxl import load_workbook
from django.db import transaction
from catalog.models import Product, Category
from .models import Receipt, ReceiptDetail
from django.utils import timezone
from inventory.services import registrar_movimiento, StockError
# from inventory.models import Movement
# from warehouse.models import Container
# from inventory.models import ProductContainer


class ReceiptImportError(Exception):
    pass


class ReceiptApprovalError(Exception):
    pass


def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def buscar_columna(encabezados, aliases):
    for alias in aliases:
        if alias in encabezados:
            return encabezados.index(alias)
    return None


def usuario_es_administrador(usuario):
    return usuario.is_superuser or getattr(usuario, "rol", "") == "ADMINISTRADOR"


@transaction.atomic
def importar_recepcion_desde_excel(*, archivo, proveedor, numero_documento, fecha_documento, usuario):
    try:
        wb = load_workbook(archivo)
        ws = wb.active
    except Exception as e:
        raise ReceiptImportError(f"No fue posible leer el archivo Excel: {e}")

    header_row = None
    encabezados = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = [normalizar_texto(v).lower() for v in row]

        if "codigo_barra" in row_values and "nombre" in row_values:
            header_row = i
            encabezados = row_values
            break

    if not header_row or not encabezados:
        raise ReceiptImportError(
            "No se encontraron encabezados válidos en el archivo Excel.")

    idx_sku = buscar_columna(encabezados, ["sku"])
    idx_codigo_barra = buscar_columna(
        encabezados, ["codigo_barra", "código_barra", "codigo barra"])
    idx_nombre = buscar_columna(
        encabezados, ["nombre", "descripcion", "descripción", "producto"])
    idx_cantidad = buscar_columna(
        encabezados, ["cantidad", "cant.", "cant", "qty"])

    faltantes = []
    if idx_sku is None:
        faltantes.append("sku")
    if idx_codigo_barra is None:
        faltantes.append("codigo_barra")
    if idx_nombre is None:
        faltantes.append("nombre")
    if idx_cantidad is None:
        faltantes.append("cantidad")

    if faltantes:
        raise ReceiptImportError(
            f"Faltan columnas obligatorias en el Excel: {', '.join(faltantes)}"
        )

    archivo.seek(0)

    recepcion = Receipt.objects.create(
        proveedor=proveedor,
        numero_documento=numero_documento,
        fecha_documento=fecha_documento,
        archivo=archivo,
        estado=Receipt.Status.BORRADOR,
        creado_por=usuario,
    )

    filas_importadas = 0
    errores = []

    for nro_fila, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        valores = [cell.value for cell in row]

        sku = normalizar_texto(
            valores[idx_sku]) if idx_sku < len(valores) else ""
        codigo_barra = normalizar_texto(
            valores[idx_codigo_barra]) if idx_codigo_barra < len(valores) else ""
        nombre = normalizar_texto(
            valores[idx_nombre]) if idx_nombre < len(valores) else ""
        cantidad_valor = valores[idx_cantidad] if idx_cantidad < len(
            valores) else None

        if not sku and not codigo_barra and not nombre:
            continue

        try:
            cantidad = int(cantidad_valor)
            if cantidad <= 0:
                raise ValueError
        except Exception:
            errores.append(f"Fila {nro_fila}: cantidad inválida.")
            continue

        producto = None

        if codigo_barra:
            producto = Product.objects.filter(
                codigo_barra=codigo_barra).first()

        if not producto and sku:
            producto = Product.objects.filter(sku=sku).first()

        ReceiptDetail.objects.create(
            recepcion=recepcion,
            producto=producto,
            sku=sku,
            codigo_barra=codigo_barra,
            nombre=nombre,
            cantidad_esperada=cantidad,
            cantidad_recibida=0,
            observacion=""
        )
        filas_importadas += 1

        if not producto:
            errores.append(
                f"Fila {nro_fila}: producto no existe en catálogo. Se importó como pendiente de revisión."
            )

    if filas_importadas == 0:
        raise ReceiptImportError(
            "No se importó ninguna fila válida. Revisa el archivo.")

    return recepcion, filas_importadas, errores


@transaction.atomic
def aprobar_recepcion(*, recepcion, usuario):
    if not usuario_es_administrador(usuario):
        raise ReceiptApprovalError(
            "Solo el Administrador puede aprobar la recepción.")

    if recepcion.estado != Receipt.Status.PENDIENTE_APROBACION:
        raise ReceiptApprovalError(
            "Solo se puede aprobar una recepción en estado Pendiente de aprobación.")

    detalles = recepcion.detalles.select_related("producto").all()
    if not detalles.exists():
        raise ReceiptApprovalError(
            "La recepción no tiene líneas para aprobar.")

    categoria_default, _ = Category.objects.get_or_create(
        nombre="Sin categoría")
    hubo_productos_aprobados = False

    for detalle in detalles:
        cantidad_aprobada = detalle.cantidad_recibida
        if cantidad_aprobada <= 0:
            continue

        producto = detalle.producto
        if not producto:
            producto = Product.objects.create(
                codigo_barra=detalle.codigo_barra,
                sku=detalle.sku,
                nombre=detalle.nombre or "Producto sin nombre",
                categoria=categoria_default,
                stock_minimo=5,
                stock_actual=0,
            )
            detalle.producto = producto
            detalle.save(update_fields=["producto"])

        try:
            # 🔹 Aquí usamos la lógica centralizada
            registrar_movimiento(
                producto_id=producto.id,
                tipo="ENTRADA",
                cantidad=cantidad_aprobada,
                usuario=usuario
            )
            hubo_productos_aprobados = True
        except StockError as e:
            raise ReceiptApprovalError(f"Error al registrar ENTRADA: {e}")

    if not hubo_productos_aprobados:
        raise ReceiptApprovalError("No hay cantidades recibidas para aprobar.")

    recepcion.estado = Receipt.Status.APROBADA
    recepcion.aprobado_por = usuario
    recepcion.fecha_aprobacion = timezone.now()
    recepcion.save(
        update_fields=["estado", "aprobado_por", "fecha_aprobacion"])

    return recepcion
